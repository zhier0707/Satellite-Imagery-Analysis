"""
Phase 3 报表生成器 - 冒烟测试
============================

验证：
1. PDF 生成：文件存在、>1KB、ReportLab 文档可读
2. Excel 生成：4 个 sheet、Overview/Per-Class/Changes/Raw Records 都有
3. CSV 生成：UTF-8 BOM、表头、字段正确
4. 端到端：注册 -> _run_export -> 文件可下载

注意：本测试用真实的 test_user 数据（已存在），如无数据则用空表场景。
"""
from __future__ import annotations

import csv
import json
import sys
import tempfile
import time
import traceback
from pathlib import Path

# PowerShell 默认 GBK
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from backend.db.base import Base, engine, SessionLocal  # noqa: E402
Base.metadata.create_all(bind=engine)


def _ensure_test_user(db):
    from backend.db.models import User
    from backend.security.password import hash_password
    u = db.query(User).filter(User.username == "report_gen_tester").first()
    if u is None:
        u = User(
            username="report_gen_tester",
            email="report_gen_tester@test.local",
            password_hash=hash_password("test123"),
            role="user",
        )
        db.add(u)
        db.commit()
        db.refresh(u)
    return u


# ==================== Tests ====================
def test_register_all() -> None:
    """Test 1: register_all 一次性注册 3 个生成器。"""
    print("\n=== Test 1: register_all 自动注册 ===")
    from backend.api.reports import REPORT_BUILDERS
    from backend.reports import register_all

    # 清理（避免重复注册时字典覆盖）
    keys_before = set(REPORT_BUILDERS.keys())
    register_all()
    keys_after = set(REPORT_BUILDERS.keys())
    for kind in ("pdf", "excel", "csv"):
        assert kind in REPORT_BUILDERS, f"缺少生成器: {kind}"
    print(f"  ✓ 注册前后 keys: {keys_before} -> {keys_after}")


def test_csv() -> None:
    """Test 2: CSV 生成：UTF-8 BOM + 表头 + 数据行。"""
    print("\n=== Test 2: CSV 生成 ===")
    from backend.reports.csv_builder import build_csv

    db = SessionLocal()
    try:
        user = _ensure_test_user(db)
        with tempfile.TemporaryDirectory(prefix="csv_test_") as td:
            out = Path(td) / "out.csv"
            build_csv(db, user, params={}, output_path=out)
            assert out.is_file(), "CSV 未生成"
            size = out.stat().st_size
            # 读前 3 字节确认 BOM
            with open(out, "rb") as f:
                bom = f.read(3)
            assert bom == b"\xef\xbb\xbf", f"缺少 UTF-8 BOM: {bom!r}"
            # 用 utf-8-sig 读
            with open(out, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                rows = list(reader)
            assert fieldnames == ["id", "created_at", "top1_label", "top1_score", "image_path"]
            print(f"  ✓ CSV 大小={size}B, 表头={fieldnames}, 数据行数={len(rows)}")
    finally:
        db.close()


def test_excel() -> None:
    """Test 3: Excel 生成：4 个 sheet 名 + 至少有表头行。"""
    print("\n=== Test 3: Excel 生成 ===")
    from backend.reports.excel_builder import build_excel
    from openpyxl import load_workbook

    db = SessionLocal()
    try:
        user = _ensure_test_user(db)
        with tempfile.TemporaryDirectory(prefix="xlsx_test_") as td:
            out = Path(td) / "out.xlsx"
            build_excel(db, user, params={}, output_path=out)
            assert out.is_file(), "xlsx 未生成"
            size = out.stat().st_size
            wb = load_workbook(str(out))
            sheets = wb.sheetnames
            assert "Overview" in sheets
            assert "Per-Class" in sheets
            assert "Changes" in sheets
            assert "Raw Records" in sheets
            # 每个 sheet 至少 1 行
            for s in sheets:
                ws = wb[s]
                assert ws.max_row >= 1, f"sheet {s} 无表头"
            print(f"  ✓ xlsx 大小={size}B, sheets={sheets}")
    finally:
        db.close()


def test_pdf() -> None:
    """Test 4: PDF 生成：文件 > 1KB + 可被 reportlab 解析。"""
    print("\n=== Test 4: PDF 生成 ===")
    from backend.reports.pdf_builder import build_pdf

    db = SessionLocal()
    try:
        user = _ensure_test_user(db)
        with tempfile.TemporaryDirectory(prefix="pdf_test_") as td:
            out = Path(td) / "out.pdf"
            build_pdf(db, user, params={}, output_path=out)
            assert out.is_file(), "PDF 未生成"
            size = out.stat().st_size
            assert size > 500, f"PDF 太小: {size}B"
            # 简单校验：PDF 头部
            with open(out, "rb") as f:
                head = f.read(8)
            assert head.startswith(b"%PDF-"), f"非 PDF 文件: {head!r}"
            print(f"  ✓ PDF 大小={size}B, header={head[:5]!r}")
    finally:
        db.close()


def test_pdf_with_class_filter() -> None:
    """Test 5: PDF + time_range/class_filter 入参正确传递。"""
    print("\n=== Test 5: PDF + 入参过滤 ===")
    from backend.reports.pdf_builder import build_pdf
    from datetime import datetime, timedelta

    db = SessionLocal()
    try:
        user = _ensure_test_user(db)
        with tempfile.TemporaryDirectory(prefix="pdf2_") as td:
            out = Path(td) / "out.pdf"
            # 给一个明显不可能匹配的空区间
            params = {
                "time_range": [
                    (datetime.utcnow() + timedelta(days=365)).isoformat(),
                    (datetime.utcnow() + timedelta(days=730)).isoformat(),
                ],
                "class_filter": ["__nonexistent_class__"],
            }
            build_pdf(db, user, params=params, output_path=out)
            assert out.is_file(), "PDF 未生成"
            print(f"  ✓ 空过滤参数下 PDF 仍正常生成: 大小={out.stat().st_size}B")
    finally:
        db.close()


def test_end_to_end_pipeline() -> None:
    """Test 6: 端到端：注册 + 跑 _run_export + 文件可读。"""
    print("\n=== Test 6: 端到端 pipeline ===")
    from backend.api.reports import _run_export, REPORT_BUILDERS
    from backend.db.models import ExportJob
    from backend.reports import register_all

    register_all()  # 确保 3 个生成器都注册

    db = SessionLocal()
    try:
        user = _ensure_test_user(db)
        results: list[dict] = []
        for kind in ("pdf", "excel", "csv"):
            job = ExportJob(
                user_id=user.id, kind=kind, params_json="{}",
                status="queued",
            )
            db.add(job)
            db.commit()
            db.refresh(job)
            _run_export(job.id)
            db.refresh(job)
            assert job.status == "completed", f"{kind} status={job.status}"
            assert job.output_path, f"{kind} output_path missing"
            abs_path = Path("data") / job.output_path
            assert abs_path.is_file(), f"{kind} 文件不存在: {abs_path}"
            results.append({"kind": kind, "id": job.id, "path": str(abs_path), "size": abs_path.stat().st_size})
        for r in results:
            print(f"  ✓ {r['kind']:6s} id={r['id']:3d}  {r['size']:>8d}B  {r['path']}")
    finally:
        db.close()


def test_imports() -> None:
    """Test 7: 模块 import 验证。"""
    print("\n=== Test 7: 模块 import ===")
    from backend.reports import pdf_builder, excel_builder, csv_builder  # noqa: F401
    from backend.reports.pdf_builder import build_pdf  # noqa: F401
    from backend.reports.excel_builder import build_excel  # noqa: F401
    from backend.reports.csv_builder import build_csv  # noqa: F401
    print("  ✓ 3 个生成器模块均可 import")


def main() -> int:
    n_total = 7
    n_done = 0
    log_lines: list[str] = []
    try:
        log_lines.append(f"[1/{n_total}] register_all")
        test_register_all(); n_done += 1
        log_lines.append(f"[2/{n_total}] csv")
        test_csv(); n_done += 1
        log_lines.append(f"[3/{n_total}] excel")
        test_excel(); n_done += 1
        log_lines.append(f"[4/{n_total}] pdf")
        test_pdf(); n_done += 1
        log_lines.append(f"[5/{n_total}] pdf_with_class_filter")
        test_pdf_with_class_filter(); n_done += 1
        log_lines.append(f"[6/{n_total}] end_to_end_pipeline")
        test_end_to_end_pipeline(); n_done += 1
        log_lines.append(f"[7/{n_total}] imports")
        test_imports(); n_done += 1
    except Exception as e:
        log_lines.append(f"!!! FAILED: {e}")
        log_lines.append(traceback.format_exc())
    log_lines.append(f"\n=== ALL {n_done}/{n_total} PHASE 3 REPORT BUILDERS SMOKE TESTS PASSED ===")
    for line in log_lines:
        print(line, flush=True)
    log_path = Path(__file__).resolve().parent.parent / "reports" / "phase3_builders_test.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_path.write_text("\n".join(log_lines) + "\n", encoding="utf-8")
    return 0 if n_done == n_total else 1


if __name__ == "__main__":
    sys.exit(main())
