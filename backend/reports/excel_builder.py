"""
Excel 报表生成器
================

调用入口：`build_excel(db, user, params, output_path)`

4 个 sheet：
- Overview：总记录数、时间范围、Top-3 类别
- Per-Class：class_name, count, percent
- Changes：该用户的全部 change_jobs
- Raw Records：classify_records 前 500 条

依赖：openpyxl
"""
from __future__ import annotations

import logging
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from sqlalchemy import and_
from sqlalchemy.orm import Session

log = logging.getLogger(__name__)


# ==================== 数据查询 ====================
def _fetch_classify_records(
    db: Session, user_id: int,
    time_range: Optional[list[str]] = None,
    class_filter: Optional[list[str]] = None,
) -> list:
    from backend.db.models import ClassifyRecord
    q = db.query(ClassifyRecord).filter(ClassifyRecord.user_id == user_id)
    if time_range and len(time_range) == 2:
        try:
            t0 = datetime.fromisoformat(time_range[0])
            t1 = datetime.fromisoformat(time_range[1])
            q = q.filter(and_(
                ClassifyRecord.created_at >= t0,
                ClassifyRecord.created_at <= t1,
            ))
        except ValueError as e:
            log.warning("invalid time_range %s: %s", time_range, e)
    rows = q.order_by(ClassifyRecord.created_at.desc()).all()
    if class_filter:
        rows = [r for r in rows if r.top1_label in class_filter]
    return rows


def _fetch_change_jobs(db: Session, user_id: int) -> list:
    from backend.db.models import ChangeJob
    return (
        db.query(ChangeJob)
        .filter(ChangeJob.user_id == user_id)
        .order_by(ChangeJob.created_at.desc())
        .all()
    )


# ==================== 主入口 ====================
def build_excel(
    db: Session,
    user,
    params: dict[str, Any],
    output_path: Path,
) -> None:
    """生成 xlsx 报表，写到 output_path。"""
    from openpyxl import Workbook

    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = _fetch_classify_records(
        db, user.id,
        time_range=params.get("time_range"),
        class_filter=params.get("class_filter"),
    )
    changes = _fetch_change_jobs(db, user.id)
    class_counter = Counter(r.top1_label for r in rows)
    total = len(rows)

    wb = Workbook()

    # ============== Overview ==============
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Field", "Value"])
    ws.append(["User", user.username])
    ws.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws.append(["Time Range", str(params.get("time_range") or "All")])
    ws.append(["Class Filter", ", ".join(params.get("class_filter") or []) or "All"])
    ws.append(["Total Records", total])
    for i, (name, cnt) in enumerate(class_counter.most_common(3), 1):
        ws.append([f"Top-{i} Class", f"{name} ({cnt})"])

    # ============== Per-Class ==============
    ws2 = wb.create_sheet("Per-Class")
    ws2.append(["class_name", "count", "percent"])
    for name, cnt in class_counter.most_common():
        pct = (cnt / total * 100) if total else 0
        ws2.append([name, cnt, round(pct, 2)])

    # ============== Changes ==============
    ws3 = wb.create_sheet("Changes")
    ws3.append(["id", "created_at", "top1_a", "top1_b", "status"])
    for cj in changes:
        ws3.append([
            cj.id,
            cj.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            cj.top1_a,
            cj.top1_b,
            cj.status,
        ])

    # ============== Raw Records ==============
    ws4 = wb.create_sheet("Raw Records")
    ws4.append(["id", "created_at", "top1_label", "top1_score", "duration_ms", "image_path"])
    for r in rows[:500]:
        ws4.append([
            r.id,
            r.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            r.top1_label,
            round(r.top1_score, 4),
            r.duration_ms,
            r.image_path,
        ])

    wb.save(str(output_path))
    log.info("[excel] %s 已生成（%d 条记录）", output_path, total)
